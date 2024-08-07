from django.shortcuts import render, redirect, get_object_or_404, get_list_or_404
from .forms import *
from  django.contrib.auth.models import User, Group
from django.contrib import messages
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required, permission_required
from .forms import platosform
from .models import *
from django.db.models import Sum
from django.utils import timezone
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from itertools import groupby
from django.http import JsonResponse
from django.core.mail import send_mail
from django.db.models import Count
from django.contrib.staticfiles.storage import staticfiles_storage
from .carro import Carro
from datetime import datetime
import json


def splash(request):
    return render(request,"splash.html")

def home(request):
    # 6 platos más vendidos
    top_platos = Descripción_Pedidos_Historico.objects.values('ID_Platos').annotate(total=Count('ID_Platos')).order_by('-total')[:6]

    # información adicional de cada plato
    platos_data = []
    for item in top_platos:
        plato = Platos.objects.get(ID_Plato=item['ID_Platos'])
        platos_data.append({
            'nombre': plato.Nombre,
            'imagen': plato.imagen.url if plato.imagen else None,
            'costo': plato.Costo,
            
        })

    context = {
        'platos_data': platos_data,
    }
    return render(request, 'index.html', context)



def reservamesa(request):
    fecha_hoy = timezone.now().date()  
    
    data= {"formremesa" : reservamesaForm }

    if request.method=="POST":
            formulario = reservamesaForm(data=request.POST)
            
            if formulario.is_valid():
                fecha_reserva = formulario.cleaned_data.get('fecha')
                
                if fecha_reserva > fecha_hoy:
                    formulario.save()
                    messages.success(request,"Reserva Enviada Exitosamente, Espere Confirmación")
                else:
                    messages.warning(request,"Debe Reservar con un dia de anticipación")
                    data["formremesa"] = formulario
            else:
                data["mensaje"] = "Error"
                data["formremesa"] = formulario


    return render(request, "reservamesa.html", data)

###esta page se llamaba login pero si la dejo como tal da problemas y los usuarios no se autologean 
def login_page (request):
    return render(request,"registration/login.html")

@login_required(login_url="/accounts/login")
@permission_required (['auth.view_group'], login_url="/")
def agregarplato(request):

    data = {"form_agregarplato" : platosform}


    if request.method=="POST":
            formulario = platosform(data=request.POST,files=request.FILES)

            if formulario.is_valid():
                messages.success(request, "Plato creado con exito, espere aprobación de administrador")
                formulario.save()
                 
            else:
                data["mensaje"] = "Error"
                data["form_agregarplato"] = formulario


    return render(request, "mantenedor/cocinero/agregarplato.html", data )



@login_required(login_url="/accounts/login")
@permission_required (['auth.add_user'], login_url="/")
def registro (request):

    data = {
        "mensaje":""
    }

    if request.method == "POST":
        nombre=request.POST.get("nombre")
        usuario=request.POST.get("usuario")
        apellido=request.POST.get("apellido")
        correo=request.POST.get("correo")
        password=request.POST.get("password")
        grupo1=request.POST.get("grupo")
        
        usu = User()
        usu.set_password(password)
        usu.username = usuario
        usu.email = correo
        usu.first_name = nombre
        usu.last_name = apellido
        usu.is_staff=(1)
        grupo = Group.objects.get(name=grupo1)
        print (grupo)
    
        try:
            usu.save()
            usu.groups.add(grupo)
            messages.success(request, "Usuario Creado Exitosamente")
            return redirect(to="registro")
                  
        except Exception as e:
            data["mensaje"] = f"Hubo un error: {e}"
            messages.error(request, data["mensaje"])
        

    return render(request,"registration/registro.html")



def registro_cli (request):

    data = {
        "mensaje":""
    }

    if request.method == "POST":
        nombre=request.POST.get("nombre")
        usuario=request.POST.get("usuario")
        apellido=request.POST.get("apellido")
        correo=request.POST.get("correo")
        passw=request.POST.get("password")
        
        usu = User()
        usu.set_password(passw)
        usu.username = usuario
        usu.email = correo
        usu.first_name = nombre
        usu.last_name = apellido
        grupo = Group.objects.get(name="Cliente")
        try:
            usu.save()
            usu.groups.add(grupo)
            user = authenticate(username=usuario, password=passw)
            print("Authenticated user:", user)  # Debugging statement
            if user is not None:
                login(request, user) 
                messages.success(request, "Usuario Creado Exitosamente")
                return redirect("index")
        except Exception as e:
            data["mensaje"] = f"Hubo un error: {e}"
            messages.error(request, data["mensaje"])
            
        
        

    return render(request,"registration/registro_cli.html")


############################################### No se que hace esta funcion  ##############################
@login_required(login_url="/accounts/login")
@permission_required (['auth.add_user'], login_url="/")
def registrouser (request):
    data = {
        'form' : agregarform
    }
    return render(request,"mantenedor/admin/agregaruser.html", data)




@login_required(login_url="/accounts/login")
@permission_required (['auth.add_group'], login_url="/")
def mesas (request):
     
    mesastb =  Mesa.objects.all()

    data = {
        "Mesas": mesastb,
          }


    return render(request, "mesas.html", data)



@login_required(login_url="/accounts/login")
@permission_required (['auth.add_user'], login_url="/")
def listaringredientes (request):
    peticiones = Sol_Ingredientes.objects.all()
    ingredientes = Bodega.objects.all()
    data = {
        'mis_ingredientes' : ingredientes ,
        'mis_peticiones' : peticiones
    }


    return render(request,"mantenedor/admin/listaringredientes.html", data )



@login_required(login_url="/accounts/login")
@permission_required (['auth.add_user'], login_url="/") 
def agregaringredientes(request):

    
    data = {
        "mensaje":""
    }

    if request.method == "POST":
        nombre=request.POST.get("nombre")
        costo=request.POST.get("costo")
        unidad=request.POST.get("unidad")

        ingre = Ingredientes()
        ingre.nombre=nombre
        ingre.Costo=costo
        ingre.unidad_medida=unidad
        print(ingre.nombre)
        print(ingre.Costo)
        print(ingre.unidad_medida)

       

        try:
            ingre.save()
            bodega = Bodega.objects.create(ID_Ingrediente=ingre, Cantidad=0)
            messages.success(request,"Ingrediente agregado correctamente")
            
           
        except:
            data["mensaje"] = "hubo un error"
        
        
        

    return render(request, "mantenedor/admin/agregaringrediente.html", data )




@login_required(login_url="/accounts/login")
@permission_required (['auth.view_group'], login_url="/")
def sol_ingredientes(request):

    data = {"form_solicitaringrediente" : Sol_Ingredientesform}
    bodegas = Bodega.objects.all()

    

    if request.method=="POST":
            formulario = Sol_Ingredientesform(data=request.POST)

            if formulario.is_valid():
                formulario.save()

            else:
                data["mensaje"] = "Error"
                data["form_solicitaringrediente"] = formulario


    return render(request, "mantenedor/cocinero/Solicitudingredientes.html", {'data': data, 'bodegas': bodegas} )



@login_required(login_url="/accounts/login")
@permission_required (['auth.add_user'], login_url="/")
def listarusuarios(request):

    Usuarios_local = User.objects.all()

    data = {
        "Usuarios" : Usuarios_local
    }


    return render(request,"mantenedor/admin/listarusuarios.html",data)   




@login_required(login_url="/accounts/login")
@permission_required (['auth.add_user'], login_url="/")
def menuadmin(request):

    Menu_local = Platos.objects.all()

    data = {
        "Menu" : Menu_local
    }

    return render(request,"mantenedor/admin/menuadmin.html", data)




@login_required(login_url="/accounts/login")
@permission_required (['auth.add_user'], login_url="/")
def modificarmenu(request, NombreBuscado):

    actual = datetime.now().time()
    if not (actual >= datetime.strptime("22:30", "%H:%M").time() or actual < datetime.strptime("08:00", "%H:%M").time()):
        messages.error(request, "Esta función solo se puede utilizar en horario no habil")
        return redirect("menuadmin")

    platom = get_object_or_404(Platos, Nombre=NombreBuscado)

    data = {

        'form': platosform(instance=platom)
    } 

    if request.method == "POST":
        formulario = platosform(data=request.POST, instance=platom, files=request.FILES)

        if formulario.is_valid():
            formulario.save()
            return redirect(to="menuadmin")
        else:
            data["mensaje"] = "Error"
            data["form"] = formulario

    return render(request,"mantenedor/admin/modificarmenu.html", data)



@login_required(login_url="/accounts/login")
@permission_required (['auth.add_user'], login_url="/")
def activaplato(request, NombreBuscado):

    platom = get_object_or_404(Platos, Nombre=NombreBuscado)
    platom.Disponibilidad=(1) 
    platom.save()    
    return redirect(to="menuadmin" )


@login_required(login_url="/accounts/login")
@permission_required (['auth.view_group'], login_url="/")
def estadoplatolisto(request, Id_pedido):

    platom = get_object_or_404(Descripción_Pedidos, ID=Id_pedido)
    platom.Listo=(1) 
    platom.save()    
    return redirect(to="estadopedidoc" )



@login_required(login_url="/accounts/login")
@permission_required (['auth.add_user'], login_url="/")
def desactivaplato(request, NombreBuscado):

    platom = get_object_or_404(Platos, Nombre=NombreBuscado)
    platom.Disponibilidad=(0) 
    platom.save()    
    return redirect(to="menuadmin" )



@login_required(login_url="/accounts/login")
@permission_required (['auth.add_user'], login_url="/")
def DescargarReporteExcel(request):
    ingrediente = Bodega.objects.all()

    wb = Workbook()
    ws = wb.active

    ws.append(['ID','Ingrediente','Cantidad'])

    

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for x in ingrediente:
        ws.append([x.ID_Ing_Bod,x.ID_Ingrediente.nombre, x.Cantidad, ])


    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max(max_length, 15) + 2
        ws.column_dimensions[column_letter].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Reporte_Inventario.xlsx'
    wb.save(response)
    
    return response
    for x in ingrediente :
        ws.append([x.ID_Ing_Bod, x.Cantidad,x.ID_Ingrediente.nombre])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Reporte_Inventario.xlsx'
        wb.save(response)
        return response
    
   
def menu(request):
    
    Menu_Platos = Platos.objects.filter(Disponibilidad=True).order_by('Region')

    platos_por_region = {}
    for region, MenuP in groupby(Menu_Platos, key=lambda x: x.Region):
        platos_por_region[region] = list(MenuP)

    data = {
        "platos_por_region": platos_por_region
    }

    return render(request, "menu.html", data)    



@login_required(login_url="/accounts/login")
@permission_required (['auth.add_user'], login_url="/")
def modificarIngrediente(request, NombreBuscado):

    ModificarI = get_object_or_404(Bodega, Nombre=NombreBuscado)

    data = {

        'form': BodegaForm(instance=ModificarI)
    } 

    if request.method == "POST":
        formulario = BodegaForm(data=request.POST, instance=ModificarI, files=request.FILES)

        if formulario.is_valid():
            formulario.save()
            return redirect(to="listaringredientes")
        else:
            data["mensaje"] = "Error"
            data["form"] = formulario

    return render(request,"mantenedor/admin/listaringredientes.html", data)

 

@login_required(login_url="/accounts/login")
@permission_required (['auth.add_user'], login_url="/")
def actualizar_ingrediente2(request, ingredientebusca):

    ingre = get_object_or_404(Bodega, ID_Ing_Bod=ingredientebusca)
    cantidadAgregar = request.POST.get("cantidad")
    ingre.Cantidad = int(cantidadAgregar) + ingre.Cantidad
    ingre.save()    
    return redirect(to="listaringredientes" )



@login_required(login_url="/accounts/login")
@permission_required (['auth.add_user'], login_url="/")
def descuenta_ingrediente(request, IdSolicitud):

    Soli = get_object_or_404(Sol_Ingredientes, ID_Solicitud_Ingrediente=IdSolicitud)
    cantidadRestar = Soli.Cantidad
    ingre = get_object_or_404(Bodega, ID_Ing_Bod=Soli.Ingrediente.ID_Ingrediente)
    ingre.Cantidad = ingre.Cantidad - cantidadRestar
    if ingre.Cantidad >= 0:
        ingre.save()    
        Soli.Realizado=(1)
        Soli.save()
        messages.success(request,"Solicitud aprobada")
    else:
            messages.error(request,"Sin Stock Suficiente, Considere Comprar Mas Ingredientes")
    return redirect(to="listaringredientes" )



@login_required(login_url="/accounts/login")
@permission_required (['auth.add_user'], login_url="/")
def listaragendamiento(request):

    Agendamientos = Reserva_Mesa.objects.all()
    fecha_hoy = timezone.localtime(timezone.now()).date()  

    data = {
        "Agendamientos" : Agendamientos,
         'fecha_hoy': fecha_hoy,     
    }

    return render(request,"mantenedor/admin/listar_agendamiento.html", data)


@login_required(login_url="/accounts/login")
@permission_required (['auth.view_user'], login_url="/")
def mesasparapedido (request):

    mesaspedido =  Mesa.objects.all()
    data = {
        "Mesas": mesaspedido,
          }
    return render(request, "mantenedor/garzon/mesasparapedido.html", data)

@login_required(login_url="/accounts/login")
@permission_required (['auth.view_user'], login_url="/")
def ingresopedidomesa (request,Mesa):

    Menu_local = Platos.objects.all()
    Pedidoss = Descripción_Pedidos.objects.all().filter(ID_Pedido=Mesa)
    try:
        pedido = Pedidos.objects.get(ID_Pedido=Mesa)
    except Pedidos.DoesNotExist:
        pedido = None

    total_cantidad = Pedidoss.aggregate(Sum('Costo'))['Costo__sum']
    if total_cantidad is None:
        total_cantidad = 0

    data = {
        "Menu" : Menu_local ,
        "mesa":Mesa,
        "Pedido":Pedidoss,
        "Pedi":pedido,
        "total_cantidad": total_cantidad,
        
    }
    return render (request,"mantenedor/garzon/ingresopedidomesa.html",data)


@login_required(login_url="/accounts/login")
@permission_required (['auth.view_user'], login_url="/")
def inicia_pedido (request, Mesaa):
    mesita = get_object_or_404(Mesa, ID_Mesa=Mesaa)
    mesita.Estado_Ocupado=(1)
    mesita.save()
    pedido = Pedidos.objects.create(ID_Pedido=Mesaa,Correo_Sol="ftecnofood@gmail.com",ID_Mesa=mesita,Estado="Sin Solicitud")

    return redirect("ingresopedidomesa",Mesa=Mesaa )



@login_required(login_url="/accounts/login")
@permission_required (['auth.view_user'], login_url="/")
def termina_pedido_nulo (request, Mesaa):
    
    pedido = get_object_or_404(Pedidos, ID_Pedido=Mesaa)
    pedido.delete()
    mesita = get_object_or_404(Mesa, ID_Mesa=Mesaa)
    mesita.Estado_Ocupado=(0)
    mesita.save()
    return redirect("ingresopedidomesa",Mesa=Mesaa )

@login_required(login_url="/accounts/login")
@permission_required (['auth.view_user'], login_url="/")
def termina_pedido (request, Mesaa):
    pedido = get_object_or_404(Pedidos, ID_Pedido=Mesaa)
    pedidos = get_list_or_404(Descripción_Pedidos, ID_Pedido=pedido)
    for pedidoss in pedidos:
        if not pedidoss.Sol_cocina:
            messages.error(request, "No se puede terminar el servicio. Algunos platos aún no han sido solicitados o eliminados.")
            return redirect("ingresopedidomesa",Mesa=Mesaa ) 

    valor=0
    mesita = get_object_or_404(Mesa, ID_Mesa=Mesaa)
    mesita.Estado_Ocupado=(0)
    mesita.save()
    pedidos = get_list_or_404(Descripción_Pedidos, ID_Pedido=Mesaa)
    mes = get_object_or_404(Mesa, ID_Mesa=Mesaa)
    for X in pedidos:
        valor = X.Costo + valor
    boleta=Boletas.objects.create(ID_Mesa=mes,Costo_Total=valor)
    bol=boleta.ID_Boleta
    for Y in pedidos:
        pedi_histo=Descripción_Pedidos_Historico.objects.create(ID_Boleta=boleta,ID_Platos=Y.ID_Platos,Costo=Y.Costo)
    pedido.delete()
    return redirect("boleta",Boleta=bol)


@login_required(login_url="/accounts/login")
@permission_required (['auth.view_user'], login_url="/")
def boleta (request, Boleta):

    bol = get_object_or_404(Boletas, ID_Boleta=Boleta)
    hist = get_list_or_404(Descripción_Pedidos_Historico, ID_Boleta=Boleta)
    data = {
        'mi_boleta' : bol ,
        'mis_peticiones' : hist
    }


    return render (request,"mantenedor/garzon/boleta.html",data)


@login_required(login_url="/accounts/login")
@permission_required (['auth.view_user'], login_url="/")
def eliminapedidolista (request, Mesaa, pedido):
    
    pedidoelimina = get_object_or_404(Descripción_Pedidos, ID=pedido)
    pedidoelimina.delete()
    return redirect("ingresopedidomesa",Mesa=Mesaa )


@login_required(login_url="/accounts/login")
@permission_required (['auth.view_user'], login_url="/")
def enviapedidococina (request, Mesaa):
    
    pedidos = get_list_or_404(Descripción_Pedidos, ID_Pedido=Mesaa)
    for pedido in pedidos:
        pedido.Sol_cocina = (1)
        pedido.save()
    return redirect("ingresopedidomesa",Mesa=Mesaa )


@login_required(login_url="/accounts/login")
@permission_required (['auth.view_user'], login_url="/")
def pedidolista (request, Mesaa, Plato):
    try:
        pedido = Pedidos.objects.get(ID_Pedido=Mesaa)
    except Pedidos.DoesNotExist:
        messages.error(request,"No se a iniciado Servicio")
        return redirect("ingresopedidomesa",Mesa=Mesaa )

    platom = get_object_or_404(Platos, ID_Plato=Plato)
    pedido= get_object_or_404(Pedidos, ID_Pedido=Mesaa)
    solicitud = Descripción_Pedidos.objects.create(ID_Pedido=pedido,ID_Platos=platom,Costo=platom.Costo)

    return redirect("ingresopedidomesa",Mesa=Mesaa )




@login_required(login_url="/accounts/login")
@permission_required (['auth.add_user'], login_url="/")
def actualizar_agendamiento(request, reserva):     
    Reserva = get_object_or_404(Reserva_Mesa, ID_reserva=reserva)
    mesa = request.POST.get("mesa")
    anterior = get_object_or_404(Mesa, ID_Mesa=Reserva.ID_Mesa.ID_Mesa)
    nueva = get_object_or_404(Mesa, ID_Mesa=mesa)

    Reserva.ID_Mesa = nueva

    if nueva.Estado_Reservado == 0:
        anterior.Estado_Reservado = 0
        Reserva.save()
        nueva.Estado_Reservado=(1)
        nueva.save()
        anterior.save()
        messages.success(request, "Agendamiento Realizado con Exito...Enviando Correo de Confirmación al Cliente. ")
        
        mensaje_correo = f"""
        Agendamiento Realizado Con Exito 

        Detalles de la reserva:
        Nombre: {Reserva.nombre}
        Correo: {Reserva.correo}
        Fecha: {Reserva.fecha}
        Hora: {Reserva.hora}
        Cantidad de comensales: {Reserva.cantidad_comensales}
        Mesa: {nueva.Nombre_Mesa}
        
        Gracias Por Preferirnos , Equipo TecnoFood

       'ftecnofood@gmail.com'
        """

        send_mail(
            'Agendamiento Realizado' ,
             mensaje_correo,
            'ftecnofood@gmail.com',
            [Reserva.correo], 
            fail_silently=False,
        )
    elif nueva.ID_Mesa == 0:
        anterior.Estado_Reservado=(0)
        Reserva.save()
        anterior.save()
    else:
        messages.error(request, "Mesa ya reservada")

    return redirect(to="listaragendamiento") 


@login_required(login_url="/accounts/login")
@permission_required (['auth.add_group'], login_url="/")
def pedidosingresadoscocina(request):

    
    mesaspedido =  Mesa.objects.all()
    pedidos =  Descripción_Pedidos.objects.all()
    reserv = Reserva_Mesa.objects.all()
    fecha_hoy = timezone.localtime(timezone.now()).date()

    data = {
        "Mesas": mesaspedido,
        "pedidos": pedidos,
        "reservas": reserv,
        "fecha_hoy":fecha_hoy
          }


    return render(request,"mantenedor/garzon/pedidosingresadoscocina.html",data)



@login_required(login_url="/accounts/login")
@permission_required (['auth.view_user'], login_url="/")
def estadopedidog (request):
     
    pedidos =  Descripción_Pedidos.objects.all()

    data = {
        "pedidos": pedidos,
          }


    return render(request, "mantenedor/garzon/estadopedidog.html", data)


@login_required(login_url="/accounts/login")
@permission_required (['auth.add_group'], login_url="/")
def estadopedidoc (request):
     
    pedidos =  Descripción_Pedidos.objects.all()

    data = {
        "pedidos": pedidos,
          }

    return render(request, "mantenedor/cocinero/estadopedidoc.html", data)


@login_required(login_url="/accounts/login")
@permission_required (['auth.view_user'], login_url="/")
def pago (request, Boleta) :

    bol = get_object_or_404(Boletas, ID_Boleta=Boleta)
    data = {
        "mi_boleta": bol
          }

    return render(request,"mantenedor/garzon/pago.html",data)


@login_required(login_url="/accounts/login")
@permission_required (['auth.view_user'], login_url="/")
def pagado (request, Boleta) :

    bol = get_object_or_404(Boletas, ID_Boleta=Boleta)
    bol.Pagado=(1)
    bol.save()


    return redirect("boleta",Boleta=Boleta )




@login_required(login_url="/accounts/login")
def funciones (request, Id_user) :
 
    user = User.objects.get(id=Id_user) 
    groups = user.groups.all()  
    grupos = groups.first().id  

    data = {
        "usuario": Id_user,
        "grupo": grupos
    }
        
    return render(request,"funciones.html", data)

###Funciones del Carrito ####

@login_required(login_url="/accounts/login")
def carrito(request, id_ped):


    pedidop= get_object_or_404(Pedidos, ID_Pedido=id_ped)
    context = {

        "pedido": pedidop


    }

    return render(request, 'Carrito/carrito.html', context)

@login_required(login_url="/accounts/login")
def agregar_plato (request, plato_id):
    
    carro=Carro(request)

    plato=Platos.objects.get(ID_Plato=plato_id)
    
    carro.agregar(plato=plato)
    

    return redirect("menupedidoonline")

@login_required(login_url="/accounts/login")
def eliminar_plato (request, plato_id):
    
    carro=Carro(request)

    plato=Platos.objects.get(ID_Plato=plato_id)
    
    carro.eliminar(plato=plato)

    return redirect("menupedidoonline")

@login_required(login_url="/accounts/login")
def restar_plato (request, plato_id):
    
    carro=Carro(request)

    plato=Platos.objects.get(ID_Plato=plato_id)
    
    carro.restar_plato(plato=plato)

    return redirect("menupedidoonline")

@login_required(login_url="/accounts/login")
def limpiar_carro (request):
    
    carro=Carro(request)
    carro.limpiar_carro()
    
    return redirect("menupedidoonline")

def eliminarP (request, id_ped):
    
    eliminarpedido = get_object_or_404(Pedidos, ID_Pedido=id_ped)
    eliminarpedido.delete()
    
    return redirect("menupedidoonline")

@login_required(login_url="/accounts/login")
@permission_required (['auth.add_user'], login_url="/")
def limpiarreserva(request):

    actual = datetime.now().time()
    if not (actual >= datetime.strptime("22:00", "%H:%M").time() or actual < datetime.strptime("08:00", "%H:%M").time()):
        messages.error(request, "Esta función solo se puede utilizar en horario no habil")
        return redirect("listaragendamiento")
    try:
        mesas_reservadas = get_list_or_404(Mesa, Estado_Reservado=1)
        for x in mesas_reservadas:
            x.Estado_Reservado = 0
            x.save()
    except :
        messages.error(request,"No se tienen mesas reservadas")
        return redirect("listaragendamiento" )

    
    return redirect(to="listaragendamiento")


@login_required(login_url="/accounts/login")
@permission_required (['auth.add_user'], login_url="/")
def eliminar_colaborador(request, usuario):
    colaborador = get_object_or_404(User, username=usuario)

    colaborador.delete()
    messages.success(request,"Colaborador eliminado correctamente")
    return redirect(to="listarusuarios")


@login_required(login_url="/accounts/login")
@permission_required (['auth.add_user'], login_url="/")
def modificar_colaborador(request, usuario):

    colaborador = get_object_or_404(User, username=usuario)

    data = {
        "form": AgregaruserForm(instance=colaborador)
    }

    if request.method == 'POST':
        formulario = AgregaruserForm(data=request.POST, instance=colaborador)
        if formulario.is_valid():
            formulario.save()
            return redirect(to="listarusuarios")
        else:
            data["mensaje"] = "Hubo un error"
            data["form"] =  formulario


    return render(request, "mantenedor/admin/modificar_colaborador.html", data)

##########         Codigo del carrito de compras           ######################


def lista_pedidos_online(request):
    
    return render (request, "Carrito/boletaonline.html")


def inicia_pedido_online (request, email):
    mesita = get_object_or_404(Mesa, ID_Mesa=0)
    pedido = Pedidos.objects.create(Correo_Sol=email,ID_Mesa=mesita,Estado="Sin Solicitud")
    id_pe = pedido.ID_Pedido
    return redirect("carrito", id_ped=id_pe)


def pedido_online_cocina(request, id_ped):
   
    carro = Carro(request)
    mesap = get_object_or_404(Mesa, ID_Mesa=0)

    for key, item in carro.carro.items():

        cant = int(item["cantidad"]) + 1
        while True:
            cant -= 1
            platop = get_object_or_404(Platos, ID_Plato=item["plato_id"])
            pedidop= get_object_or_404(Pedidos, ID_Pedido=id_ped)
            solicitud = Descripción_Pedidos.objects.create(ID_Pedido=pedidop,ID_Platos=platop,Costo=platop.Costo)
            if int(cant) >> 1:
                continue
            else:
                break 
    
    
    pedido = get_object_or_404(Pedidos, ID_Pedido=id_ped)
    valort=0
    pedidot = get_list_or_404(Descripción_Pedidos, ID_Pedido=pedido)
    for X in pedidot:
        valort = X.Costo + valort
        X.Sol_cocina=(1)
        X.save()
    boleta=Boletas.objects.create(ID_Mesa=mesap,Costo_Total=valort)
    bol=boleta.ID_Boleta
    for Y in pedidot:
        pedi_histo=Descripción_Pedidos_Historico.objects.create(ID_Boleta=boleta,ID_Platos=Y.ID_Platos,Costo=Y.Costo)
    
    bol = get_object_or_404(Boletas, ID_Boleta=bol)
    hist = get_list_or_404(Descripción_Pedidos_Historico, ID_Boleta=bol)
    data = {
        'mi_boleta' : bol ,
        'mis_peticiones' : hist
        
    }
    
    messages.success(request, "Pedido realizado con exito")

    return render (request,"Carrito/boletaonline.html",data)            


def pago_online (request, Boleta) :

    bol = get_object_or_404(Boletas, ID_Boleta=Boleta)
    data = {
        "mi_boleta": bol
          }

    return render(request,"Carrito/pagoonline.html",data)


def pagado_online (request, Boleta) :

    bol = get_object_or_404(Boletas, ID_Boleta=Boleta)
    bol.Pagado=(1)
    bol.save()

 
    return redirect("boleta_online",Boleta=Boleta )


def boleta_online (request, Boleta):

    bol = get_object_or_404(Boletas, ID_Boleta=Boleta)
    hist = get_list_or_404(Descripción_Pedidos_Historico, ID_Boleta=Boleta)
    data = {
        'mi_boleta' : bol ,
        'mis_peticiones' : hist
    }


    return render (request,"Carrito/boletaonline.html",data)


def correo_boleta(request, email ):
        carro = Carro(request)
        mensaje= """
        Gracias por su compra

        Detalles de su pedido:
        """

        for key, item in carro.carro.items():
        
            mensaje += f"""
            - {item["nombre"]}  X  {item["cantidad"]}
                                        
            """

        mensaje += f"""
            Total: ${sum(int(item["precio"]) for item in carro.carro.values())}

        Gracias por preferirnos,
        Equipo TecnoFood
        

        'ftecnofood@gmail.com'

        """

        
        messages.success(request, "Correo enviado con exito , espere confirmación de pedido listo")
           
        send_mail(
            'Pedido Realizado',
            mensaje,
            'ftecnofood@gmail.com',
            [email],
            fail_silently=False,
        )
        
        carro.limpiar_carro()

        ped = get_list_or_404(Pedidos, Correo_Sol=email)
        ultimo_pedido = ped[-1]
        ultimo_pedido.Estado = "Solicitado a Cocina"
        ultimo_pedido.save()


        return redirect(to="menupedidoonline")



@login_required(login_url="/accounts/login")
@permission_required (['auth.view_user'], login_url="/")
def estadopedidoonline (request):

    mesap = get_object_or_404(Mesa, ID_Mesa=0)
    pedidos =  Pedidos.objects.all().filter(ID_Mesa=mesap)

    data = {
        "pedidos": pedidos,
          }

    return render(request, "mantenedor/garzon/Estadopedidoonline.html", data)


@login_required(login_url="/accounts/login")
@permission_required (['auth.view_user'], login_url="/")
def solicitar_retiro_pedido (request, pedido):

    pedidoelimina = get_object_or_404(Pedidos, ID_Pedido=pedido)
    descripciones = Descripción_Pedidos.objects.filter(ID_Pedido=pedidoelimina)

    if all(descripcion.Listo for descripcion in descripciones):    
        mensaje= """

        Estimado/a Cliente,

        Nos complace informarle que su pedido ya está listo para ser retirado.

        Agradecemos su preferencia y esperamos servirle nuevamente en el futuro.

        Atentamente,
        Equipo TecnoFood

        Correo de contacto: ftecnofood@gmail.com



        """

        
        messages.success(request, "Correo enviado con exito , espere por retiro de cliente")
           
        send_mail(
                'Pedido Realizado',
                mensaje,
                'ftecnofood@gmail.com',
                [pedidoelimina.Correo_Sol],
                fail_silently=False,
            )
        
        pedidoelimina.Estado=("Esperar retiro cliente")
        pedidoelimina.save()
    else:
        messages.error(request, "El pedido no se encuentra listo para el retiro, espere confirmación de cocina")

    return redirect("estadopedidoonline")
    
@login_required(login_url="/accounts/login")
@permission_required (['auth.view_user'], login_url="/")
def terminapedidoonline (request, pedido):
    
    pedidoelimina = get_object_or_404(Pedidos, ID_Pedido=pedido)
    pedidoelimina.delete()
    messages.success(request, "Pedido entregado al cliente")

    return redirect("estadopedidoonline")    

@login_required(login_url="/accounts/login")
def menupedidoonline(request):
    Menu_Platos = Platos.objects.filter(Disponibilidad=True).order_by('Region')

    platos_por_region = {}
    for region, MenuP in groupby(Menu_Platos, key=lambda x: x.Region):
        platos_por_region[region] = list(MenuP)

    data = {
        "platos_por_region": platos_por_region
    }
    return render(request,"Carrito/menupedidoonline.html", data)
   
def reportes(request):
    return render(request,"mantenedor/admin/reportes.html")

